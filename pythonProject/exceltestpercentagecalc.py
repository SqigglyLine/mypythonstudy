import openpyxl as xl
from openpyxl.chart import (
    LineChart,
    Reference,
)


class ScoreImprovement:
    def __init__(self, filename):
        self.filename = filename

    def percentCalcNewScores(self):
        # This function finds percentages - displays them, then adds improvments to the scores
        wb = xl.load_workbook(self.filename)
        sheet = wb['Sheet1']
        # Finds the percentage
        for cells in range(2, sheet.max_row + 1):
            cell = sheet.cell(cells, 2)
            initalpercentage = cell.value / 130 * 100
            initalpercentage_cell = sheet.cell(cells, 4)
            initalpercentage_cell.value = initalpercentage
        # Improves the scores
        for cells in range(2, sheet.max_row + 1):
            cell = sheet.cell(cells, 4)
            if cell.value < 94:
                newpercent = cell.value + 5
                newpercent_cell = sheet.cell(cells, 5)
                newpercent_cell.value = newpercent
            elif cell.value == 100:
                sheet.cell(cells, 5).value = 100
            else:
                newpercent = cell.value + cell.value * 0.05
                newpercent_cell = sheet.cell(cells, 5)
                newpercent_cell.value = newpercent
            # This if statement is needed to make sure the percentages don't go over 100%
            if newpercent_cell.value > 100:
                newpercent_cell.value = 100
        # This finds the new scores
        for cells in range(2, sheet.max_row + 1):
            cell = sheet.cell(cells, 5)
            newscore = cell.value / 100 * 130
            newscore_cell = sheet.cell(cells, 3)
            newscore_cell.value = newscore

        values = Reference(sheet,
                           min_row=2,
                           max_row=sheet.max_row,
                           min_col=4,
                           max_col=5)
        may_test_scores = LineChart()
        may_test_scores.add_data(values)

        may_test_scores.title = 'Test Scores'
        may_test_scores.y_axis.title = 'Score'
        may_test_scores.x_axis.title = 'Students'

        sheet.add_chart(may_test_scores, 'g2')

        wb.save(self.filename)

    def highestScorer(self):  # ***WORK ON LATER***
        wb = xl.load_workbook(self.filename)
        sheet = wb['Sheet1']

        for cells in range(2, sheet.max_row):
            scores = []
            scores += sheet.cell(cells, 5).value

        max = scores[0]
        for number in scores:
            if number > max:
                max = number

        sheet.cell(7, 2).value = max

        wb.save(self.filename)


scores_May = ScoreImprovement('exceltest.xlsx')
scores_May.percentCalcNewScores()
